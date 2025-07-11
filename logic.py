"""
表格处理逻辑模块
实现销售出库单和即时库存表的同步功能
"""

import pandas as pd
import re
import os
from typing import Dict, List, Tuple, Optional, Callable
from utils import load_excel_file, save_excel_file, clean_dataframe, get_column_by_name, create_backup_file
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class StockSyncProcessor:
    """库存同步处理器"""
    
    def __init__(self, progress_callback: Optional[Callable] = None):
        self.progress_callback = progress_callback
        self.sales_df = None
        self.stock_df = None
        self.material_mapping = {}  # 物料编码映射 {旧料号: 新料号}
        self.modified_cells = []  # 记录修改的单元格位置
        self.mapping_config_file = 'material_mapping.json'
        
        # 自动加载保存的映射配置
        self._load_mapping_config()
        
    def set_progress_callback(self, callback: Callable):
        """设置进度回调函数"""
        self.progress_callback = callback
        
    def _update_progress(self, message: str):
        """更新进度"""
        if self.progress_callback:
            self.progress_callback(message)
    
    def load_sales_file(self, file_path: str) -> str:
        """
        加载销售出库单文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            错误信息，空字符串表示成功
        """
        try:
            self._update_progress("正在加载销售出库单...")
            df, error = load_excel_file(file_path, dtype=str)
            
            if error:
                return error
            
            # 不进行数据清理，保持原始结构（包括空行）
            self.sales_df = df
            self.sales_file_path = file_path
            
            # 验证必要的列是否存在（按列位置检查）
            required_columns = {
                'DZ': 129,  # 第130列 - 物料编码
                'EC': 132,  # 第133列 - 辅助属性1
                'ED': 133,  # 第134列 - 辅助属性2
                'FF': 161,  # 第162列 - 批号#主档
                'FG': 162,  # 第163列 - 批号#手工
                'GJ': 191,  # 第192列 - (明细信息)仓库#名称
                'HA': 208   # 第209列 - 销售数量
            }
            
            max_col_needed = max(required_columns.values())
            if len(self.sales_df.columns) <= max_col_needed:
                return f"销售出库单列数不足，需要至少 {max_col_needed + 1} 列，实际只有 {len(self.sales_df.columns)} 列"
            
            # 为方便后续处理，给需要的列添加别名
            for alias, col_idx in required_columns.items():
                if col_idx < len(self.sales_df.columns):
                    self.sales_df[alias] = self.sales_df.iloc[:, col_idx]
            
            self._update_progress(f"销售出库单加载成功，共 {len(self.sales_df)} 行数据")
            return ""
            
        except Exception as e:
            return f"加载销售出库单失败: {str(e)}"
    
    def load_stock_file(self, file_path: str) -> str:
        """
        加载即时库存表文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            错误信息，空字符串表示成功
        """
        try:
            self._update_progress("正在加载即时库存表...")
            df, error = load_excel_file(file_path)
            
            if error:
                return error
            
            # 不进行数据清理，保持原始结构（包括空行）
            self.stock_df = df
            
            # 验证必要的列是否存在（按列位置检查）
            required_columns = {
                'A': 0,   # 第1列 - 物料编码
                'G': 6,   # 第7列 - 仓库名称
                'H': 7,   # 第8列 - 批号
                'K': 10,  # 第11列 - 库存数量
                'D': 3,   # 第4列 - 辅助属性
                'E': 4,   # 第5列 - 辅助属性
                'F': 5    # 第6列 - 辅助属性
            }
            
            max_col_needed = max(required_columns.values())
            if len(self.stock_df.columns) <= max_col_needed:
                return f"即时库存表列数不足，需要至少 {max_col_needed + 1} 列，实际只有 {len(self.stock_df.columns)} 列"
            
            # 为方便后续处理，给需要的列添加别名
            for alias, col_idx in required_columns.items():
                if col_idx < len(self.stock_df.columns):
                    self.stock_df[alias] = self.stock_df.iloc[:, col_idx]
            
            self._update_progress(f"即时库存表加载成功，共 {len(self.stock_df)} 行数据")
            return ""
            
        except Exception as e:
            return f"加载即时库存表失败: {str(e)}"
    
    def set_material_mapping(self, old_code: str, new_code: str) -> str:
        """
        设置物料编码映射
        
        Args:
            old_code: 旧料号
            new_code: 新料号
            
        Returns:
            错误信息，空字符串表示成功
        """
        if not old_code or not new_code:
            return "物料编码不能为空"



        # 简单验证编码格式
        if not self._validate_material_code(old_code) or not self._validate_material_code(new_code):
            return "物料编码格式不正确，应为 x.xx.x.xx.xx.xxx 格式"

        self.material_mapping[old_code] = new_code
        # 保存映射配置到文件
        self._save_mapping_config()
        return ""
    
    def set_material_mappings(self, mappings: List[tuple]) -> str:
        """
        批量设置物料编码映射
        
        Args:
            mappings: 映射列表，每个元素为 (old_code, new_code) 元组
            
        Returns:
            错误信息，空字符串表示成功
        """
        if not mappings:
            return "映射列表不能为空"
        
        errors = []
        for old_code, new_code in mappings:
            if not old_code or not new_code:
                errors.append(f"物料编码不能为空: {old_code} -> {new_code}")
                continue
                
            # 简单验证编码格式
            if not self._validate_material_code(old_code) or not self._validate_material_code(new_code):
                errors.append(f"物料编码格式不正确: {old_code} -> {new_code}")
                continue
            


            self.material_mapping[old_code] = new_code
        
        if errors:
            return "; ".join(errors)
        
        # 保存映射配置到文件
        self._save_mapping_config()
        return ""
    
    def clear_material_mappings(self):
        """清空所有物料编码映射"""
        self.material_mapping = {}
        self._save_mapping_config()
    
    def get_material_mappings(self) -> dict:
        """获取当前的物料编码映射"""
        return self.material_mapping.copy()
    
    def _load_mapping_config(self):
        """从JSON文件加载映射配置"""
        try:
            import json
            import os
            
            if os.path.exists(self.mapping_config_file):
                with open(self.mapping_config_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # 对键和值进行规范化，避免因空白字符导致匹配失败
                    self.material_mapping = {
                        self._normalize_material_code(k): self._normalize_material_code(v)
                        for k, v in data.items()
                    }
                    self._update_progress(f"已加载 {len(self.material_mapping)} 个物料编码映射")
        except Exception as e:
            # 如果加载失败，保持空映射
            self.material_mapping = {}
            self._update_progress(f"加载映射配置失败: {e}")
    
    def _save_mapping_config(self):
        """保存映射配置到JSON文件"""
        try:
            import json

            with open(self.mapping_config_file, 'w', encoding='utf-8') as f:
                json.dump(self.material_mapping, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self._update_progress(f"保存映射配置失败: {e}")

    def _normalize_material_code(self, code: str) -> str:
        """
        仅清理物料编码中的空格和不可见字符，不移除小数点，也不转换为数字
        """
        if pd.isna(code):
            return ""
        return str(code).strip().replace("\u200b", "").replace("\xa0", "")

    def _validate_material_code(self, code: str) -> bool:
        """验证物料编码格式"""
        if not code or not isinstance(code, str):
            return False
        
        parts = code.split('.')
        if len(parts) != 6:
            return False
        
        for part in parts:
            if not part.isdigit():
                return False
        
        return True
    
    def _log_section(self, title):
        self._update_progress("\n" + "="*20 + f" {title} " + "="*20 + "\n")

    def process_synchronization(self) -> str:
        """
        执行同步处理
        
        Returns:
            错误信息，空字符串表示成功
        """
        try:
            if self.sales_df is None:
                return "请先加载销售出库单"
            
            if self.stock_df is None:
                return "请先加载即时库存表"
            
            if not self.material_mapping:
                # 如果没有映射，只执行后续步骤可能无意义，但暂不强制要求
                self._update_progress("警告：未设置物料编码映射，将跳过料号替换步骤。")

            # 清空上次的修改记录
            self.modified_cells = []
            
            # 开始处理
            self._update_progress("开始处理数据同步...")
            
            # 1. 替换物料编码
            self._log_section("物料编码替换")
            if self.material_mapping:
                self._replace_material_codes()
            
            # 2. 同步辅助属性
            self._log_section("辅助属性同步")
            self._sync_auxiliary_attributes()

            # 3. 批号同步（如有需要）
            self._log_section("批号同步")
            self._sync_batch_numbers()

            # 保存文件并高亮修改内容
            self._save_with_highlights()
            
            self._update_progress("数据同步完成！")
            return ""
            
        except Exception as e:
            return f"同步处理失败: {str(e)}"
    
    def _replace_material_codes(self):
        """根据映射表批量替换物料编码并记录详细日志"""
        self._update_progress("开始替换物料编码（字符串匹配）...")

        success_total = 0
        failed_total = 0
        fail_reasons = []

        # 1. 先全部转字符串（防止Excel读取为数值）
        self.sales_df['DZ'] = self.sales_df['DZ'].astype(str)

        for old_code, new_code in self.material_mapping.items():
            self._update_progress(
                f"现在开始使用字符串替换旧料号 {old_code} 成新料号 {new_code}"
            )
            old_norm = self._normalize_material_code(old_code)
            new_norm = self._normalize_material_code(new_code)

            if not old_norm or old_norm == new_norm:
                self._update_progress(
                    f"跳过 {old_code} -> {new_code}: 新旧料号相同或为空"
                )
                failed_total += 1
                continue

            # 2. 用布尔掩码找目标行
            mask_old = (self.sales_df['DZ'].apply(self._normalize_material_code) == old_norm) & (self.sales_df.index >= 1)
            mask_new = (self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm) & (self.sales_df.index >= 1)
            before_count = mask_old.sum()
            new_before = mask_new.sum()

            if before_count == 0:
                msg = f"未找到需要替换的物料编码 {old_code}"
                self._update_progress(msg)
                fail_reasons.append(msg)
                failed_total += 1
                continue

            # 替换前，打印被选中的行号和原始值
            sample_rows = self.sales_df.loc[mask_old, :].head(5)
            self._update_progress(f"将要替换的前5行索引和值：\n{sample_rows[['DZ']].to_dict(orient='index')}")

            # 3. 替换为新料号（字符串直接赋值）
            self.sales_df.loc[mask_old, 'DZ'] = new_code

            # 同步更新原始物料编码列并记录修改单元格
            original_dz_idx = 129  # 物料编码原始列索引（0-based）
            if original_dz_idx < len(self.sales_df.columns):
                changed_indices = self.sales_df.index[mask_old]
                for row_idx in changed_indices:
                    self.sales_df.iat[row_idx, original_dz_idx] = new_code
                    self.modified_cells.append((row_idx, original_dz_idx))
                self._update_progress(f"本次同步写回原始物料编码列的行索引: {list(changed_indices)}")

            # 替换后，打印被选中的行号和新值
            sample_rows_after = self.sales_df.loc[mask_old, :].head(5)
            self._update_progress(f"替换后前5行索引和值：\n{sample_rows_after[['DZ']].to_dict(orient='index')}")

            # 4. 统计替换后情况
            old_remaining = (
                self.sales_df['DZ'].apply(self._normalize_material_code) == old_norm
            ).sum()
            after_count = (
                self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm
            ).sum()
            replaced_count = after_count - new_before

            success_total += replaced_count
            failed_total += old_remaining
            self._update_progress(
                f"已将 {old_code} 替换为 {new_code}，应替换 {before_count} 行，实际成功 {replaced_count} 行，剩余 {old_remaining} 行，新料号共 {after_count} 行（其中原有 {new_before} 行）"
            )

        self._update_progress(
            f"物料编码替换完成，总成功 {success_total} 行，剩余未替换 {failed_total} 行"
        )

        if fail_reasons:
            for reason in fail_reasons:
                self._update_progress(f"失败原因: {reason}")


    def _sync_auxiliary_attributes(self):
        """
        只同步那些已经替换成新料号的行的辅助属性，且用新料号+仓库查库存表。
        日志详细说明每一步。
        """
        self._update_progress("开始同步辅助属性（仅处理新料号行）...")

        # 构建库存表查找字典（新料号+仓库）
        stock_aux_map = {}
        temp_stock_df = self.stock_df.iloc[1:].copy()
        temp_stock_df['K_num'] = pd.to_numeric(temp_stock_df['K'], errors='coerce').fillna(0)
        for _, row in temp_stock_df.iterrows():
            material_code = self._normalize_material_code(row['A'])
            warehouse = str(row['G']).strip()
            if material_code and warehouse:
                current_stock_qty = row['K_num']
                if (material_code, warehouse) not in stock_aux_map or current_stock_qty > stock_aux_map.get((material_code, warehouse), {}).get('K', -1):
                    stock_aux_map[(material_code, warehouse)] = {
                        'E': '' if pd.isna(row['E']) else str(row['E']),
                        'F': '' if pd.isna(row['F']) else str(row['F']),
                        'K': current_stock_qty
                    }

        ec_col_idx = 132
        ed_col_idx = 133

        # 保证顺序与映射表一致
        for new_code in list(dict.fromkeys(self.material_mapping.values())):
            new_norm = self._normalize_material_code(new_code)
            # 先筛选出所有新料号的行
            sales_df_new = self.sales_df.iloc[1:][self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm]
            warehouse_list = sales_df_new['GJ'].dropna().astype(str).str.strip().unique()
            self._update_progress(f"新料号 {new_norm} 共涉及 {len(warehouse_list)} 个仓库: {list(warehouse_list)}")
            for warehouse_str in warehouse_list:
                mask = (
                    (self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm) &
                    (self.sales_df['GJ'].astype(str).str.strip() == warehouse_str) &
                    (self.sales_df.index >= 1)
                )
                indices = self.sales_df.index[mask]
                if len(indices) == 0:
                    continue

                self._update_progress(f"处理新料号: '{new_norm}' + 仓库: '{warehouse_str}'，共 {len(indices)} 行。")

                # 销售出库表原本的EC/ED属性
                old_ec_vals = self.sales_df.loc[indices, 'EC'].unique()
                old_ed_vals = self.sales_df.loc[indices, 'ED'].unique()
                self._update_progress(f"  -> 销售出库表原本EC属性: {list(old_ec_vals)}，ED属性: {list(old_ed_vals)}")

                # 即时库存表中该料号+仓库的辅助属性
                match = stock_aux_map.get((new_norm, warehouse_str))
                if match:
                    aux_e = match['E']
                    aux_f = match['F']
                    self._update_progress(f"  -> 即时库存表中该料号+仓库的E列属性: '{aux_e}'，F属性: '{aux_f}'")
                    self._update_progress(f"  -> 将销售出库表中{len(indices)}行的EC属性从{list(old_ec_vals)}改为'{aux_e}'，ED属性从{list(old_ed_vals)}改为'{aux_f}'")
                    self.sales_df.loc[indices, 'EC'] = aux_e
                    if ec_col_idx < len(self.sales_df.columns):
                        for idx in indices:
                            self.sales_df.iat[idx, ec_col_idx] = aux_e
                            if (idx, ec_col_idx) not in self.modified_cells:
                                self.modified_cells.append((idx, ec_col_idx))
                    self.sales_df.loc[indices, 'ED'] = aux_f
                    if ed_col_idx < len(self.sales_df.columns):
                        for idx in indices:
                            self.sales_df.iat[idx, ed_col_idx] = aux_f
                            if (idx, ed_col_idx) not in self.modified_cells:
                                self.modified_cells.append((idx, ed_col_idx))
                else:
                    self._update_progress(f"  -> 警告: 在库存表中未找到新料号 '{new_norm}' 在仓库 '{warehouse_str}' 的记录，跳过此组合的辅助属性同步。")

        self._update_progress("辅助属性同步完成。")


    def _sync_batch_numbers(self):
        """
        同步销售出库单中的批次号，按库存分配，详细日志说明每一步。
        """
        self._update_progress("开始同步批号（仅处理新料号行，按库存分配）...")

        temp_stock_df = self.stock_df.iloc[1:].copy()
        temp_stock_df['K_num'] = pd.to_numeric(temp_stock_df['K'], errors='coerce').fillna(0)

        ff_col_idx = 161
        fg_col_idx = 162

        for new_code in list(dict.fromkeys(self.material_mapping.values())):
            new_norm = self._normalize_material_code(new_code)
            sales_df_new = self.sales_df.iloc[1:][self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm]
            warehouse_list = sales_df_new['GJ'].dropna().astype(str).str.strip().unique()
            self._update_progress(f"新料号 {new_norm} 共涉及 {len(warehouse_list)} 个仓库: {list(warehouse_list)}")
            for warehouse_str in warehouse_list:
                mask = (
                    (self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm) &
                    (self.sales_df['GJ'].astype(str).str.strip() == warehouse_str) &
                    (self.sales_df.index >= 1)
                )
                indices = self.sales_df.index[mask]
                if len(indices) == 0:
                    continue

                self._update_progress(f"处理新料号: '{new_norm}' + 仓库: '{warehouse_str}'，共 {len(indices)} 行。")

                # 销售出库表原本的批号FF/FG属性
                old_ff_vals = self.sales_df.loc[indices, 'FF'].unique()
                old_fg_vals = self.sales_df.loc[indices, 'FG'].unique()
                self._update_progress(f"  -> 销售出库表原本FF(批号主档)属性: {list(old_ff_vals)}，FG(批号手工)属性: {list(old_fg_vals)}")

                # 1. 获取即时库存表中该料号+仓库下所有批号及可用库存
                stock_rows = temp_stock_df[
                    (temp_stock_df['A'].apply(self._normalize_material_code) == new_norm) &
                    (temp_stock_df['G'].astype(str).str.strip() == warehouse_str)
                ]
                if stock_rows.empty:
                    self._update_progress(f"  -> 警告: 在库存表中未找到新料号 '{new_norm}' 在仓库 '{warehouse_str}' 的记录，跳过此组合的批号同步。")
                    continue

                # 按库存量K从大到小排序
                stock_rows = stock_rows.sort_values(by='K_num', ascending=False)
                batch_info = []
                for _, row in stock_rows.iterrows():
                    batch = '' if pd.isna(row['H']) else str(row['H'])
                    qty = int(row['K_num'])
                    batch_info.append({'batch': batch, 'qty': qty})
                batch_info_str = ', '.join([f"{b['batch']}({b['qty']})" for b in batch_info])
                self._update_progress(f"  -> 即时库存表批号分布: [{batch_info_str}]")

                # 2. 分配批号
                sales_indices = list(indices)
                allocated = 0
                for b in batch_info:
                    if not sales_indices:
                        break
                    assign_count = min(b['qty'], len(sales_indices))
                    if assign_count == 0:
                        continue
                    self._update_progress(f"    -> 批号 '{b['batch']}' 可用库存 {b['qty']}，分配 {assign_count} 行。")
                    for i in range(assign_count):
                        idx = sales_indices.pop(0)
                        self.sales_df.at[idx, 'FF'] = b['batch']
                        self.sales_df.at[idx, 'FG'] = b['batch']
                        if ff_col_idx < len(self.sales_df.columns) and fg_col_idx < len(self.sales_df.columns):
                            self.sales_df.iat[idx, ff_col_idx] = b['batch']
                            self.sales_df.iat[idx, fg_col_idx] = b['batch']
                            if (idx, ff_col_idx) not in self.modified_cells:
                                self.modified_cells.append((idx, ff_col_idx))
                            if (idx, fg_col_idx) not in self.modified_cells:
                                self.modified_cells.append((idx, fg_col_idx))
                        allocated += 1
                if sales_indices:
                    self._update_progress(f"  -> 警告: 库存不足，尚有 {len(sales_indices)} 行未能分配批号！")
                else:
                    # 分配完成后，输出FF、FG列的新值
                    new_ff_vals = self.sales_df.loc[indices, 'FF'].unique()
                    new_fg_vals = self.sales_df.loc[indices, 'FG'].unique()
                    self._update_progress(f"  -> 批号分配完成，共分配 {allocated} 行。销售出库表FF列新值: {list(new_ff_vals)}，FG列新值: {list(new_fg_vals)}")

        self._update_progress("批号同步完成。")

    def _synchronize_by_flow(self):
        """按照给定流程同步批次号和辅助属性"""
        self._update_progress("正在同步批次和辅助属性...")

        mappings = list(self.material_mapping.items())
        total = len(mappings)
        n = 1
        processed_codes = set()

        # 收集销售出库单中出现的仓库顺序列表（跳过前两行）
        warehouse_list = []
        for w in self.sales_df['GJ'].iloc[2:]:
            if pd.notna(w) and w not in warehouse_list:
                warehouse_list.append(w)

        for old_code, new_code in mappings:
            new_norm = self._normalize_material_code(new_code)

            if new_norm in processed_codes:
                n += 1
                continue

            processed_codes.add(new_norm)
            self._update_progress(
                f"处理料号 {n}/{total}: {old_code} -> {new_code}")

            for warehouse in warehouse_list:
                sales_rows = self.sales_df[
                    (self.sales_df.index >= 2) &
                    (self.sales_df['DZ'].apply(self._normalize_material_code) == new_norm) &
                    (self.sales_df['GJ'].astype(str).str.strip() == str(warehouse).strip())
                ]
                if sales_rows.empty:
                    continue

                stock_rows = self.stock_df[
                    (self.stock_df.index >= 1) &
                    (self.stock_df['A'].apply(self._normalize_material_code) == new_norm) &
                    (self.stock_df['G'].astype(str).str.strip() == str(warehouse).strip())
                ]

                if stock_rows.empty:
                    self._update_progress(
                        f"警告: 库存表中没有找到仓库 {warehouse} 的料号 {new_code}")
                    continue

                stock_row = stock_rows.sort_values(by='K', ascending=False).iloc[0]
                batch = stock_row['H']
                aux_e = stock_row['E']
                aux_f = stock_row['F']

                indices = sales_rows.index

                self.sales_df.loc[indices, 'FF'] = batch
                self.sales_df.loc[indices, 'FG'] = batch
                self.sales_df.loc[indices, 'EC'] = aux_e
                self.sales_df.loc[indices, 'ED'] = aux_f

                ff_idx, fg_idx, ec_idx, ed_idx = 161, 162, 132, 133
                if ff_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[indices, ff_idx] = batch
                if fg_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[indices, fg_idx] = batch
                if ec_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[indices, ec_idx] = aux_e
                if ed_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[indices, ed_idx] = aux_f

                for idx in indices:
                    self.modified_cells.extend([
                        (idx, ff_idx),
                        (idx, fg_idx),
                        (idx, ec_idx),
                        (idx, ed_idx)
                    ])

            n += 1
    
    def _process_warehouses(self):
        """处理所有仓库的数据"""
        self._update_progress("正在处理仓库数据...")
        
        # 获取所有仓库列表（跳过标题行和空行）
        warehouse_data = []
        for idx in range(2, len(self.sales_df)):
            # 跳过完全空行
            if self.sales_df.iloc[idx].isna().all():
                continue
                
            warehouse = self.sales_df.at[idx, 'GJ']
            if pd.notna(warehouse) and warehouse != '':
                warehouse_data.append(warehouse)
        
        warehouses = list(set(warehouse_data))  # 去重
        
        for warehouse in warehouses:
            self._update_progress(f"正在处理仓库: {warehouse}")
            self._process_warehouse_data(warehouse)
    
    def _process_warehouse_data(self, warehouse: str):
        """处理单个仓库的数据"""
        # 获取该仓库的所有新料号（跳过标题行和空行）
        material_codes = []
        
        for idx in range(2, len(self.sales_df)):
            # 跳过完全空行
            if self.sales_df.iloc[idx].isna().all():
                continue

            row_warehouse = self.sales_df.at[idx, 'GJ']
            row_material = self._normalize_material_code(self.sales_df.at[idx, 'DZ'])
            
            if (pd.notna(row_warehouse) and row_warehouse == warehouse and
                pd.notna(row_material) and row_material != ''):
                material_codes.append(row_material)
        
        # 去重处理每个物料
        unique_materials = list(set(material_codes))
        
        for material_code in unique_materials:
            self._process_material_in_warehouse(material_code, warehouse)
    
    def _process_material_in_warehouse(self, material_code: str, warehouse: str):
        """处理仓库中的具体物料"""
        material_code = self._normalize_material_code(material_code)


        # 从即时库存表中获取该物料在该仓库的库存信息（跳过标题行和空行）
        stock_rows = []
        
        for idx in range(1, len(self.stock_df)):  # 即时库存表从第2行开始
            # 跳过完全空行
            if self.stock_df.iloc[idx].isna().all():
                continue

            row_material = self._normalize_material_code(self.stock_df.at[idx, 'A'])
            row_warehouse = str(self.stock_df.at[idx, 'G']).strip()
            
            if (pd.notna(row_material) and row_material == material_code and
                pd.notna(row_warehouse) and row_warehouse == str(warehouse)):
                stock_rows.append(idx)
        
        if not stock_rows:
            self._update_progress(f"警告: 在库存表中未找到物料 {material_code} 在仓库 {warehouse} 的信息")
            return
        
        # 获取销售出库单中对应的行（跳过标题行和空行）
        sales_rows = []
        
        for idx in range(2, len(self.sales_df)):  # 销售出库单从第3行开始
            # 跳过完全空行
            if self.sales_df.iloc[idx].isna().all():
                continue

            row_material = self._normalize_material_code(self.sales_df.at[idx, 'DZ'])
            row_warehouse = str(self.sales_df.at[idx, 'GJ']).strip()
            
            if (pd.notna(row_material) and row_material == material_code and
                pd.notna(row_warehouse) and row_warehouse == str(warehouse)):
                sales_rows.append(idx)
        
        if not sales_rows:
            return
        
        # 按库存量分配批次号
        self._allocate_batch_numbers(sales_rows, stock_rows, material_code, warehouse)
        
        # 更新辅助属性
        self._update_auxiliary_attributes(sales_rows, stock_rows, material_code, warehouse)
    
    def _allocate_batch_numbers(self, sales_row_indices: list, stock_row_indices: list,
                               material_code: str, warehouse: str):
        """分配批次号"""

        # 按行数计算需要分配的数量
        total_sales_qty = len(sales_row_indices)
        if total_sales_qty <= 0:
            return
        
        # 收集库存批次信息
        batch_info = {}
        for idx in stock_row_indices:
            batch_num = self.stock_df.at[idx, 'H']
            batch_qty = self.stock_df.at[idx, 'K']
            
            if pd.notna(batch_num) and pd.notna(batch_qty):
                try:
                    qty = float(batch_qty)
                    if batch_num not in batch_info:
                        batch_info[batch_num] = {
                            'quantity': 0,
                            'auxiliary_attrs': {
                                'D': self.stock_df.at[idx, 'D'],
                                'E': self.stock_df.at[idx, 'E'],
                                'F': self.stock_df.at[idx, 'F']
                            }
                        }
                    batch_info[batch_num]['quantity'] += qty
                except:
                    pass
        
        # 按库存量分配批次号
        allocated_qty = 0
        batch_allocation = []
        
        for batch_num, info in batch_info.items():
            if allocated_qty >= total_sales_qty:
                break
                
            batch_stock = int(info['quantity'])

            # 计算本批次可分配的行数
            remaining_qty = total_sales_qty - allocated_qty
            allocated_batch_qty = min(batch_stock, remaining_qty)
            
            if allocated_batch_qty > 0:
                batch_allocation.append({
                    'batch_num': batch_num,
                    'quantity': allocated_batch_qty,
                    'auxiliary_attrs': info['auxiliary_attrs']
                })
                
                allocated_qty += allocated_batch_qty
        
        # 应用分配结果到销售出库单
        self._apply_batch_allocation(sales_row_indices, batch_allocation)
    
    def _apply_batch_allocation(self, sales_row_indices: list, batch_allocation: List[Dict]):
        """应用批次分配结果"""
        row_idx = 0
        
        for allocation in batch_allocation:
            batch_num = allocation['batch_num']
            quantity = int(allocation['quantity'])
            auxiliary_attrs = allocation['auxiliary_attrs']

            rows_to_allocate = quantity

            while rows_to_allocate > 0 and row_idx < len(sales_row_indices):
                actual_idx = sales_row_indices[row_idx]

                # 更新批次号
                self.sales_df.at[actual_idx, 'FF'] = batch_num
                self.sales_df.at[actual_idx, 'FG'] = batch_num

                # 同时更新原始列
                ff_col_idx = 161  # FF列的实际位置
                fg_col_idx = 162  # FG列的实际位置

                if ff_col_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[actual_idx, ff_col_idx] = batch_num
                if fg_col_idx < len(self.sales_df.columns):
                    self.sales_df.iloc[actual_idx, fg_col_idx] = batch_num

                # 记录修改的单元格
                self.modified_cells.extend([
                    (actual_idx, ff_col_idx),
                    (actual_idx, fg_col_idx)
                ])

                rows_to_allocate -= 1
                row_idx += 1
    
    def _update_auxiliary_attributes(self, sales_row_indices: list, stock_row_indices: list,
                                   material_code: str, warehouse: str):
        """更新辅助属性"""
        if not stock_row_indices:
            return
        
        # 获取库存表中的辅助属性（使用第一个库存行的属性）
        first_stock_idx = stock_row_indices[0]
        aux_attrs = {
            'D': self.stock_df.at[first_stock_idx, 'D'],
            'E': self.stock_df.at[first_stock_idx, 'E'],
            'F': self.stock_df.at[first_stock_idx, 'F']
        }
        
        # 更新销售出库单中的辅助属性
        # 根据需要可以在这里添加辅助属性的同步逻辑
        # 目前题目中没有明确销售出库单的辅助属性列位置，所以暂时跳过
        pass
    
    def _save_with_highlights(self):
        """
        将修改保存到新文件中，以绕过原文件的单元格保护问题。
        新文件将包含所有修改和高亮。
        """
        self._update_progress("正在生成新文件以保存修改...")

        # 1. 定义新文件名
        original_path = self.sales_file_path
        path_without_ext, ext = os.path.splitext(original_path)
        new_file_path = f"{path_without_ext}_modified{ext}"
        
        self._update_progress(f"所有修改将被保存到新文件: {new_file_path}")

        # 2. 再次用 openpyxl 加载原始文件到内存
        try:
            wb = load_workbook(original_path)
            ws = wb.active
        except Exception as e:
            self._update_progress(f"错误：无法加载原始Excel文件 '{original_path}' 进行保存。 {e}")
            return

        # 3. 设置高亮样式
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

        # 4. 将所有在内存中（DataFrame）的修改应用到 openpyxl 对象上
        for row_idx, col_idx in self.modified_cells:
            # 转换列索引为 openpyxl 使用的1-based格式
            openpyxl_col_idx = self._get_column_index(col_idx)
            
            # 获取修改后的新值
            new_value = self.sales_df.iat[row_idx, col_idx]

            # DataFrame 的行索引对应 Excel 的物理行号 (DataFrame索引 + 2)
            excel_row = row_idx + 2

            # 更新单元格的值并应用高亮
            cell = ws.cell(row=excel_row, column=openpyxl_col_idx)
            cell.value = new_value
            cell.fill = red_fill
            
        # 5. 保存到新文件
        try:
            wb.save(new_file_path)
            self._update_progress(f"成功！已将修改保存到新文件: {new_file_path}")
        except Exception as e:
            self._update_progress(f"错误：保存新文件失败！ {e}")

    def _get_column_index(self, column_index: int) -> int:
        """将0-based列索引转换为1-based列索引（openpyxl使用）"""
        return column_index + 1
    
    def get_warehouses_count(self, material_code: str) -> int:
        """获取指定物料的仓库数量"""
        if self.stock_df is None:
            return 0

        code = self._normalize_material_code(material_code)
        warehouses = self.stock_df[self.stock_df['A'].apply(self._normalize_material_code) == code]['G'].nunique()
        return warehouses
    
    def get_batch_info(self, material_code: str, warehouse: str) -> List[Dict]:
        """获取指定物料在指定仓库的批次信息"""
        if self.stock_df is None:
            return []


        code = self._normalize_material_code(material_code)
        batch_info = self.stock_df[
            (self.stock_df['A'].apply(self._normalize_material_code) == code) &
            (self.stock_df['G'].astype(str).str.strip() == str(warehouse))
        ]
        
        result = []
        for _, row in batch_info.iterrows():
            result.append({
                'batch_num': row['H'],
                'quantity': row['K'],
                'auxiliary_attrs': {
                    'D': row['D'],
                    'E': row['E'],
                    'F': row['F']
                }
            })
        return result
